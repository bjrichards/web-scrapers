import random
import time
from dataclasses import dataclass
from typing import Optional

import bs4
import openpyxl
import requests


# car dataclass
@dataclass
class car:
    vin: str = ""
    stock_num: Optional[str] = ""
    year: Optional[str] = ""
    title: Optional[str] = ""
    price: Optional[str] = ""
    mileage: Optional[str] = ""
    drivetrain: Optional[str] = ""
    sale_link: Optional[str] = ""

    def __str__(self):
        return f"\tVIN: {self.vin} \n\
        Stock Number: {self.stock_num} \n\
        Year: {self.year} \n\
        Title: {self.title} \n\
        Price: {self.price} \n\
        Mileage: {self.mileage} \n\
        Drivetrain: {self.drivetrain} \n\
        Sales Link: {self.sale_link}"


# Cars.com
class scraper_cars_dot_com:
    prefix: str = "https://www.cars.com/shopping/results/?"
    link_prefix: str = "https://www.cars.com"
    tags: list[str] = [
        "maximum_distance=250",  # max mileage away (250 miles)
        "makes[]=porsche",  # make of Porsche
        "models[]=porsche-911",  # model of 911
        "transmission_slugs[]=manual",  # manual transmission
        "page_size=100",  # page size of 100
        "sort=best_match_desc",  # needed sort choice for website
        "stock_type=used",  # used cars only
        "zip=89519",  # zip code of area I want to shop around
    ]

    soup: bs4.BeautifulSoup
    links: list = []
    cars: list[car] = []
    new_cars: list[car] = []

    def __init__(self) -> None:
        pass

    def run(self) -> None:
        self.retrieve_site()
        self.grab_links()
        self.get_each_car_info()

        excel_contr: excel_controller = excel_controller()
        existing_vins = excel_contr.grab_workbook_data_vin("911 For Sale")
        self.get_new_vehicles(existing_vins)

    def retrieve_site(self) -> None:
        url = self.compile_url()
        page = requests.get(url)
        self.soup = bs4.BeautifulSoup(page.content, "html.parser")

    def grab_links(self) -> None:
        for link in self.soup.find_all("a", {"class": "vehicle-card-link"}, href=True):
            if link.text:
                self.links.append(link["href"])

    def get_each_car_info(self) -> None:
        for link in self.links:
            car_link = self.link_prefix + link
            print(f"Grabbing from: {car_link}")
            car_info = self.get_car_info(car_link)
            if car_info:
                self.cars.append(car_info)
            time.sleep(random.randint(0, 3))

    def get_car_info(self, car_link) -> Optional[car]:
        result: car = car()
        cleaned_id_text: list[str] = []
        cleaned_id_attribute_text: list[str] = []

        car_page = requests.get(car_link)
        soup = bs4.BeautifulSoup(car_page.content, "html.parser")

        # Get drivetrain, vin, stock #, and mileage
        info_group = soup.find_all("dl", {"class": "fancy-description-list"})
        for i in info_group[0].find_all("dt"):
            cleaned_id_text.append(i.text)

        for i in info_group[0].find_all("dd"):
            cleaned_id_attribute_text.append(i.text)

        for i in range(0, len(cleaned_id_text)):
            id: str = cleaned_id_text[i]
            data: str = cleaned_id_attribute_text[i]
            if id == "Drivetrain":
                result.drivetrain = data
            elif id == "VIN":
                result.vin = data
            elif id == "Stock #":
                result.stock_num = data
            elif id == "Mileage":
                result.mileage = data.replace(".", "")

        # Get price
        price = soup.find("span", {"class": "primary-price"}).text  # type: ignore
        result.price = price

        # Get title
        title = soup.find("h1", {"class": "listing-title"}).text  # type: ignore
        result.title = title

        # Set sales link
        result.sale_link = car_link

        # Get Year
        result.year = title[0:5]

        if result.vin:
            return result
        else:
            return None

    def compile_url(self) -> str:
        result: str = ""

        result += self.prefix

        for tag in self.tags:
            result += tag
            result += "&"

        # remove hangin '&' on end
        result = result[:-1]

        return result

    def get_new_vehicles(self, existing_vins: list[str]) -> None:
        for car in self.cars:
            if car.vin not in existing_vins:
                self.new_cars.append(car)


class excel_controller:
    workbook_name: str = "911_for_sale.xlsx"

    def add_data_to_workbook(self, data: list[car], worksheet_name: str) -> None:
        wb = openpyxl.load_workbook(filename=self.workbook_name)

        if worksheet_name in wb.sheetnames:
            wb.remove(wb[worksheet_name])
        ws = wb.active
        ws.title = worksheet_name

        row = col = 1

        for car in data:
            ws.cell(column=col, row=row, value=car.vin)
            ws.cell(column=col + 1, row=row, value=car.year)
            ws.cell(column=col + 2, row=row, value=car.title)
            ws.cell(column=col + 3, row=row, value=car.price)
            ws.cell(column=col + 4, row=row, value=car.mileage)
            ws.cell(column=col + 5, row=row, value=car.drivetrain)
            ws.cell(column=col + 6, row=row, value=car.stock_num)
            ws.cell(column=col + 7, row=row, value=car.sale_link)

            row += 1

        wb.save(filename=self.workbook_name)

    def grab_workbook_data_vin(self, worksheet_name: str) -> list[str]:
        wb = openpyxl.load_workbook(filename=self.workbook_name)
        ws1 = wb[worksheet_name]
        result: list[str] = []
        for row in ws1.iter_rows():  # type: ignore
            if row[0]:
                result.append(row[0].value)

        return result


if __name__ == "__main__":
    cars_com: scraper_cars_dot_com = scraper_cars_dot_com()
    excel_contr: excel_controller = excel_controller()

    cars_com.run()
    excel_contr.add_data_to_workbook(cars_com.cars, "911 For Sale")
    # vins = excel_contr.grab_workbook_data_vin("911 For Sale")
    excel_contr.add_data_to_workbook(cars_com.new_cars, "New 911s For Sale")
